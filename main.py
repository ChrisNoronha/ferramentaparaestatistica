import math
from openpyxl import Workbook
from ajustar_amplitude_classes import ajusta_amplitude, ajusta_classes

def ordenar_contar_e_calcular_amplitude(nome_arquivo):
    # Lê os números do arquivo e os armazena em uma lista
    numeros = []
    with open(nome_arquivo, 'r') as arquivo:
        for linha in arquivo:
            numeros_linha = linha.strip().split()
            numeros.extend(map(int, numeros_linha))

    # Ordena os números em ordem crescente
    numeros_ordenados = sorted(numeros)

    # Conta quantos números há no total
    total_numeros = len(numeros)

    # Calcula a amplitude total
    amplitude = numeros_ordenados[-1] - numeros_ordenados[0]

    # Calcula a quantidade de classes usando a fórmula de Sturges
    quantidade_classes = round(math.sqrt(total_numeros))

    # Modifica a amplitude das classes, se necessário
    amplitude_classe = round((amplitude / quantidade_classes))
    amplitude_classe = ajusta_amplitude(amplitude_classe)
    quantidade_classes = ajusta_classes(quantidade_classes)

    # Inicializa as listas para armazenar os limites e os pontos médios das classes
    limites_classes = []
    pontos_medios_classes = []

    # Define o primeiro limite inferior como o menor número no conjunto de dados
    limite_inferior = numeros_ordenados[0]

    # Calcula os limites e os pontos médios de cada classe
    for _ in range(quantidade_classes):
        limite_superior = limite_inferior + amplitude_classe
        ponto_medio = (limite_inferior + limite_superior) / 2
        limites_classes.append((limite_inferior, limite_superior))
        pontos_medios_classes.append(ponto_medio)
        limite_inferior = limite_superior

    return numeros_ordenados, total_numeros, amplitude, quantidade_classes, amplitude_classe, limites_classes, pontos_medios_classes

# Nome do arquivo contendo os números
nome_arquivo = 'numeros_rol.txt'

# Chama a função principal para calcular as estatísticas
numeros_ordenados, total_numeros, amplitude, quantidade_classes, amplitude_classe, limites_classes, pontos_medios_classes = ordenar_contar_e_calcular_amplitude(nome_arquivo)

# Conta quantos elementos estão em cada classe
contagem_por_classe = []
for limite_inf, limite_sup in limites_classes:
    contagem_classe = sum(limite_inf <= numero < limite_sup for numero in numeros_ordenados)
    contagem_por_classe.append(contagem_classe)

# Calcula a frequência acumulada
frequencia_acumulada = 0
frequencia_acumulada_por_classe = []
for contagem in contagem_por_classe:
    frequencia_acumulada += contagem
    frequencia_acumulada_por_classe.append(frequencia_acumulada)
freq_acul=0
freq_acul=frequencia_acumulada
# Calcula a frequência multiplicada pelo ponto médio de cada classe
frequencia_por_ponto_medio = [contagem * ponto_medio for contagem, ponto_medio in zip(contagem_por_classe, pontos_medios_classes)]






# Exibe os resultados
print("\nNúmeros ordenados:", numeros_ordenados)
print("\nTotal de números:", total_numeros)
print("\nAmplitude:", amplitude)
print("\nQuantidade de classes:", quantidade_classes)
print("\nAmplitude de cada classe:", amplitude_classe)
print("\nLimites de cada classe:")
for i, limite in enumerate(limites_classes, start=1):
    print(f"Classe {i}: {limite[0]} - {limite[1]}")
print("\nPonto médio de cada classe:")
for i, ponto_medio in enumerate(pontos_medios_classes, start=1):
    print(f"Classe {i}: {ponto_medio}")

print("\nContagem de elementos em cada classe (f):")
for i, (contagem, freq_acumulada) in enumerate(zip(contagem_por_classe, frequencia_acumulada_por_classe), start=1):
    print(f"Classe {i}: {contagem}  Frequência acumulada (F): {freq_acumulada}")
print(f"\nFrequência total: {frequencia_acumulada}")

print("\nFrequência multiplicada pelo ponto médio, (Pm*i):")
for i, valor in enumerate(frequencia_por_ponto_medio, start=1):
    print(f"Classe {i}: {valor}")
print(f"Total: {sum(frequencia_por_ponto_medio)}")

# Calcula a média
media = sum(frequencia_por_ponto_medio) / frequencia_acumulada

# Calcula a diferença entre ponto médio e média, ao quadrado, e multiplica pela frequência
diferencas_quadradas = [(ponto_medio - media) ** 2 * contagem for ponto_medio, contagem in zip(pontos_medios_classes, contagem_por_classe)]
variancia = sum(diferencas_quadradas) / frequencia_acumulada
desvio_padrao = math.sqrt(variancia)

# Calcula a moda
index_moda = contagem_por_classe.index(max(contagem_por_classe))
moda = pontos_medios_classes[index_moda]

# Arredonda os resultados para 5 dígitos após a vírgula
media = round(media, 5)
variancia = round(variancia, 5)
desvio_padrao = round(desvio_padrao, 5)
moda = round(moda, 5)

media_menos_ponto_medio_quadrado = [(ponto_medio-media)**2 for contagem, ponto_medio in zip(contagem_por_classe, pontos_medios_classes)]

media_menos_ponto_medio_quadrado_por_freq = [(ponto_medio-media)**2*contagem for contagem, ponto_medio in zip(contagem_por_classe, pontos_medios_classes)]

def valor_moda(moda):
    return moda
valor_moda(moda)
print(f"\nMédia: {media}")
print(f"Variância: {variancia}")
print(f"Desvio Padrão: {desvio_padrao}")
print(f"Moda: {moda}")

# Calcula a mediana
if total_numeros % 2 == 1:
    mediana_index = total_numeros // 2
    mediana = numeros_ordenados[mediana_index]
else:
    mediana_index_1 = total_numeros // 2 - 1
    mediana_index_2 = total_numeros // 2
    mediana = (numeros_ordenados[mediana_index_1] + numeros_ordenados[mediana_index_2]) / 2

mediana = round(mediana, 5)
print(f"\nMediana: {mediana}")

# Cria uma nova planilha Excel
wb = Workbook()
ws = wb.active
ws.title = "Estatísticas"

# Adiciona um cabeçalho à planilha
ws.append(["Classe", "Limite", "P Médio", "Frequência", "Frequência Ac", "Freq * PM", "PM -Media","(Freq * PM)²","(Freq * PM)²*fi"])
soma_freq_pm=sum(frequencia_por_ponto_medio)
soma_media_pm_quad=sum(media_menos_ponto_medio_quadrado)
soma_media_pm_quad_freq=sum(media_menos_ponto_medio_quadrado_por_freq)

# Preenche a planilha com os limites das classes e pontos médios, começando da célula B2
for i, (limite, ponto_medio, frequencia, frequencia_acumulada, freq_pm) in enumerate(zip(limites_classes, pontos_medios_classes, contagem_por_classe, frequencia_acumulada_por_classe, frequencia_por_ponto_medio), start=2):
    ws.cell(row=i, column=1, value=f"{i - 1}")
    ws.cell(row=i, column=2, value=f"{limite[0]} - {limite[1]}")
    ws.cell(row=i, column=3, value=ponto_medio)
    ws.cell(row=i, column=4, value=frequencia)
    ws.cell(row=i, column=5, value=frequencia_acumulada)
    ws.cell(row=i+1, column=4, value=freq_acul)
    ws.cell(row=i, column=7, value=ponto_medio - media)
    ws.cell(row=i, column=6, value=freq_pm)  
    ws.cell(row=i+1, column=6, value=soma_freq_pm) 
    ws.cell(row=i, column=8, value=(ponto_medio - media)**2)
    ws.cell(row=i, column=9, value=((ponto_medio - media)**2)*frequencia)
    ws.cell(row=i+1, column=9, value=soma_media_pm_quad_freq)
ws.cell(row=1, column=11, value="media")
ws.cell(row=2, column=11, value=media)
ws.cell(row=3, column=11, value="Variancia")
ws.cell(row=4, column=11, value=variancia)
ws.cell(row=5, column=11, value="Desvio Padrao")
ws.cell(row=6, column=11, value=desvio_padrao)
ws.cell(row=7, column=11, value="moda")
ws.cell(row=8, column=11, value=moda)
ws.cell(row=9, column=11, value="mediana")
ws.cell(row=10, column=11, value=mediana)
    
# Salva a planilha em um arquivo
wb.save("classes.xlsx")